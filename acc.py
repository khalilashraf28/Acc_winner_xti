import tkinter as tk
from tkinter import filedialog, messagebox
import pandas as pd
import threading
from tkinter.ttk import Progressbar
import numpy as np
import openpyxl
from datetime import datetime, timedelta
import os
from openpyxl.styles import PatternFill

# Function to process files
def process_files():
    try:
        # Start progress bar
        progress_bar.start(10)

        # Load data from files
        df_kpi = pd.read_excel(file_paths['Metro KPI'], engine="xlrd")
        df_perf = pd.read_excel(file_paths['Employee Performance'], engine="xlrd")
        df_tsheet = pd.read_csv(file_paths['TimeSheet'])

        # Processing logic (simplified example for demonstration)
        df_kpi = df_kpi[df_kpi['totaccessory'] > 1200]
        final_df = df_kpi.copy()
        final_df['Store ACC Sold'] = df_kpi['totaccessory']

        # Add yesterday's date column
        yesterday = datetime.now() - timedelta(days=1)
        final_df['Date'] = yesterday.strftime('%d %B %Y')

        # Map additional columns
        final_df['Market'] = final_df['Store ACC Sold'].map(
            lambda x: df_kpi.loc[df_kpi['totaccessory'] == x, 'marketid'].iloc[0] if x in df_kpi['totaccessory'].values else ""
        )
        final_df['Store ID'] = final_df['Store ACC Sold'].map(
            lambda x: df_kpi.loc[df_kpi['totaccessory'] == x, 'custno'].iloc[0] if x in df_kpi['totaccessory'].values else ""
        )
        final_df['Store'] = final_df['Store ACC Sold'].map(
            lambda x: df_kpi.loc[df_kpi['totaccessory'] == x, 'company'].iloc[0] if x in df_kpi['totaccessory'].values else ""
        )

        # Create NTID column
        final_df['NTID'] = final_df.apply(
            lambda row: get_username(row, df_perf), axis=1
        )

        # Add Employee Name
        name_lookup = df_tsheet.set_index('username')['name'].to_dict()
        lastname_lookup = df_tsheet.set_index('username')['lastname'].to_dict()
        full_name_lookup = {k: f"{name_lookup[k]} {lastname_lookup[k]}" for k in name_lookup}
        final_df['Employee name'] = final_df['NTID'].map(full_name_lookup).fillna("")

        # Add EMP ACC Sold column
        final_df['EMP ACC Sold'] = final_df.apply(
            lambda row: calculate_emp_acc_sold(row, df_perf), axis=1
        )

        # Add Payout column
        final_df['Payout'] = final_df.apply(calculate_value, axis=1)
        final_df = final_df[['Date',"Market",'Store ID','Store','NTID','Employee name','EMP ACC Sold','Payout']]
        print(final_df.head())
        # Save or update the Excel file
        save_to_excel_with_header_color(final_df, "Acc Winners(auto).xlsx")

        # Stop spinner and show success message
        progress_bar.stop()
        messagebox.showinfo("Success", "Process Completed Successfully")

    except Exception as e:
        progress_bar.stop()
        messagebox.showerror("Error", f"An error occurred: {str(e)}")

# Helper functions
def get_username(row, df_perf):
    if row['Market']:
        filtered = df_perf[
            (df_perf['company'] == row['Store']) &
            (df_perf['custno'] == row['Store ID'])
        ]
        if not filtered.empty:
            return filtered.loc[filtered['totaccessory'].idxmax(), 'username']
    return ""

def calculate_emp_acc_sold(row, df_perf):
    if row['Market']:
        filtered = df_perf[
            (df_perf['company'] == row['Store']) &
            (df_perf['custno'] == row['Store ID'])
        ]
        return filtered['totaccessory'].max() if not filtered.empty else np.nan
    return np.nan

def calculate_value(row):
    if row['Market']:
        if row['Market'] == "SOUTH FL 2" or (
            row['Store'] in ["NORTHWEST HWY", "704 JEFFERSON"] and row['Store ACC Sold'] < 1500
        ):
            return 0
        return 25
    return np.nan

def save_to_excel_with_header_color(final_df, file_path, header_color="#E49EDD"):
    if not os.path.exists(file_path):
        final_df.to_excel(file_path, index=False)
    else:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
        for _, row in final_df.iterrows():
            sheet.append(row.tolist())
        workbook.save(file_path)
    apply_header_color(file_path, header_color)

def apply_header_color(file_path, header_color):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    fill = PatternFill(start_color=header_color[1:], end_color=header_color[1:], fill_type="solid")
    for col in range(1, sheet.max_column + 1):
        sheet.cell(row=1, column=col).fill = fill
    workbook.save(file_path)

# Tkinter UI setup
def upload_file(file_key):
    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xls;*.xlsx"), ("CSV files", "*.csv")])
    if file_path:
        file_paths[file_key] = file_path
        labels[file_key].config(text=f"{file_key} Loaded: {os.path.basename(file_path)}")

def start_processing():
    if all(key in file_paths for key in ['Metro KPI', 'Employee Performance', 'TimeSheet']):
        threading.Thread(target=process_files).start()
    else:
        messagebox.showerror("Error", "Please upload all required files before proceeding.")

# Main UI
root = tk.Tk()
root.title("File Upload and Processing")
root.geometry("500x300")

file_paths = {}
labels = {}
file_keys = ['Metro KPI', 'Employee Performance', 'TimeSheet']

y_pos = 50
for key in file_keys:
    labels[key] = tk.Label(root, text=f"{key} not loaded", width=60, anchor="w")
    labels[key].place(x=20, y=y_pos)
    tk.Button(root, text="Upload", command=lambda k=key: upload_file(k)).place(x=400, y=y_pos - 5)
    y_pos += 40

tk.Button(root, text="Start Processing", command=start_processing).place(x=200, y=y_pos)
progress_bar = Progressbar(root, mode='indeterminate')
progress_bar.place(x=150, y=y_pos + 50, width=200)

root.mainloop()



